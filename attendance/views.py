from rest_framework import generics, views, viewsets
from rest_framework.decorators import api_view
from rest_framework.parsers import FileUploadParser
from rest_framework.response import Response
from rest_framework import status
from django.http import Http404
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import json

from users.models import Teacher
from .models import AttendanceStatus, TeachersAttendance
from .serializers import (AttendanceStatusSerializer, TeachersAttendanceSerializer)

class AttendanceStatusViewSet(viewsets.ModelViewSet):
	queryset = AttendanceStatus.objects.all()
	serializer_class = AttendanceStatusSerializer

class TeachersAttendanceViewSet(viewsets.ModelViewSet):
	queryset = TeachersAttendance.objects.all()
	serializer_class = TeachersAttendanceSerializer

class TeachersAttendanceListView(views.APIView):
	"""
    List all students, or create a new student.
    """
	def get(self, request, format=None):
		attendances = TeachersAttendance.objects.all()
		serializer = TeachersAttendanceSerializer(attendances, many=True)
		return Response(serializer.data)

	def post(self, request, format=None):
		serializer = TeachersAttendanceSerializer(data=request.data)
		print(request.data)
		print(serializer.is_valid())
		if serializer.is_valid():
			serializer.save()
			return Response(serializer.data, status=status.HTTP_201_CREATED)
		return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class TeachersAttendanceDetailView(views.APIView):
	def get_object(self, pk):
		try:
			return TeachersAttendance.objects.get(pk=pk)
		except TeachersAttendance.DoesNotExist:
			raise Http404

	def get(self, request, pk, format=None):
		attendance = self.get_object(pk)
		serializer = TeachersAttendanceSerializer(attendance)
		return Response(serializer.data)
		
	def put(self, request, pk, format=None):
		attendance = self.get_object(pk)
		serializer = TeachersAttendanceSerializer(attendance, data=request.data)
		if serializer.is_valid():
			serializer.save()
			return Response(serializer.data)
		return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
	
	def delete(self, request, pk, format=None):
		attendance = self.get_object(pk)
		attendance.delete()
		return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
def dailyAttendanceView(request):
	date = request.query_params.get('date')
	attendance = TeachersAttendance.objects.filter(date=date)
	
	serializer = TeachersAttendanceSerializer(attendance, many=True)
	return Response({'teacher-attendance': serializer.data})

class TeachersAttendanceBulkCreateView(generics.CreateAPIView):
	serializer_class = TeachersAttendanceSerializer
	queryset = TeachersAttendance.objects.all()

	def post(self, request):
		print(request.data)
		serializer = TeachersAttendanceSerializer(request.data)
		#serializer.is_valid()
		teachers_attendance = serializer.create(request)
		print(teachers_attendance)
		for attendance in teachers_attendance:

			serializer = TeachersAttendanceSerializer(data=attendance)
			serializer.is_valid()
			if attendance:
				continue
				#return Response(status=status.HTTP_201_CREATED)
		return Response(status=status.HTTP_400_BAD_REQUEST)

class TeachersAttendanceBulkUploadView(views.APIView):
	"""
	This uploads bulk daily teacher's attendance from an excel file
	"""

	parser_class = [FileUploadParser]
	def post(self, request, filename, format="xlsx"):
		file_obj = request.data
		xlfile = file_obj["filename"]

		print(xlfile)
		wb = load_workbook(xlfile)
		ws = wb.active
		print(ws.title)

		teachers_att = []
		for row in ws.iter_rows(min_row=2, max_col=9, max_row=31, values_only=True):
			teachers_att.append(row)
			#print(api)
			
		attendances = []
		for i in range(len(teachers_att)):
			teacher = {
				"date": f"{teachers_att[i][0]}",
				"time_in": f"{teachers_att[i][1]}",
				"time_out": f"{teachers_att[i][2]}",
				"teacher": f"{teachers_att[i][3]}",
				"status": f"{teachers_att[i][4]}",
					}
			attendances.append(teacher)
			
		for teacher in attendances:
			serializer = TeachersAttendanceSerializer(data=teacher)
			print(serializer.is_valid())
			if serializer.is_valid():
				serializer.save()
				#return Response(serializer.data, status=status.HTTP_201_CREATED)
		return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



